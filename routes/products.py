import io
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from models.product import (PRODUCT_CATEGORIES, PRODUCT_UNITS, TRANSACTION_TYPES,
                             ORDER_TYPES, ORDER_STATUSES, COUNT_STATUSES)
from services.product_service import ProductService
from datetime import datetime

products_bp = Blueprint('products', __name__, url_prefix='/products')
service = ProductService()


# ============================================================
# DANH SÁCH SẢN PHẨM
# ============================================================
@products_bp.route('/')
@login_required
def list_products():
    tenant_id = current_user.tenant_id
    search = request.args.get('search', '').strip()
    category = request.args.get('category', '')
    stock_filter = request.args.get('stock', '')
    page = request.args.get('page', 1, type=int)

    filters = {}
    if category:
        filters['category'] = category
    if stock_filter:
        filters['stock'] = stock_filter

    products = service.get_list(tenant_id, filters=filters, search=search, page=page)
    stats = service.get_list_stats(tenant_id)

    return render_template('products/list.html', products=products,
                           search=search, category=category, stock_filter=stock_filter,
                           categories=PRODUCT_CATEGORIES,
                           total_count=stats['total_count'],
                           active_count=stats['active_count'],
                           low_stock=stats['low_stock'],
                           total_value=stats['total_value'])


# ============================================================
# CRUD SẢN PHẨM
# ============================================================
@products_bp.route('/create', methods=['GET', 'POST'])
@login_required
def create():
    if request.method == 'POST':
        product = service.create_product(current_user.tenant_id, request.form, current_user.id)
        flash('Đã thêm sản phẩm mới!', 'success')
        return redirect(url_for('products.list_products'))

    return render_template('products/form.html', product=None,
                           categories=PRODUCT_CATEGORIES, units=PRODUCT_UNITS,
                           title='Thêm sản phẩm')


@products_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit(id):
    product = service.get_one(current_user.tenant_id, id)
    if request.method == 'POST':
        service.update_product(current_user.tenant_id, id, request.form)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'ok', 'message': 'Đã lưu'})

        flash('Đã cập nhật sản phẩm!', 'success')
        return redirect(url_for('products.list_products'))

    return render_template('products/form.html', product=product,
                           categories=PRODUCT_CATEGORIES, units=PRODUCT_UNITS,
                           title='Sửa sản phẩm')


@products_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete(id):
    service.delete(current_user.tenant_id, id)
    flash('Đã xóa sản phẩm.', 'success')
    return redirect(url_for('products.list_products'))


# ============================================================
# GIAO DỊCH ĐƠN LẺ
# ============================================================
@products_bp.route('/<int:id>/stock', methods=['POST'])
@login_required
def adjust_stock(id):
    product, txn = service.adjust_stock(
        current_user.tenant_id, id, request.form, current_user.id
    )
    flash(f'{dict(TRANSACTION_TYPES).get(txn.type, txn.type)}: '
          f'{abs(txn.quantity)} {product.unit} — {product.name}', 'success')
    return redirect(url_for('products.stock_history', id=product.id))


@products_bp.route('/<int:id>/history')
@login_required
def stock_history(id):
    product = service.get_one(current_user.tenant_id, id)
    page = request.args.get('page', 1, type=int)
    transactions = service.get_stock_history(id, page=page)
    return render_template('products/history.html', product=product,
                           transactions=transactions, transaction_types=TRANSACTION_TYPES)


@products_bp.route('/transactions')
@login_required
def all_transactions():
    txn_type = request.args.get('type', '')
    page = request.args.get('page', 1, type=int)
    transactions = service.get_all_transactions(
        current_user.tenant_id, txn_type=txn_type or None, page=page
    )
    return render_template('products/transactions.html',
                           transactions=transactions, txn_type=txn_type,
                           transaction_types=TRANSACTION_TYPES)


# ============================================================
# PHIẾU NHẬP/XUẤT KHO (StockOrder)
# ============================================================
@products_bp.route('/orders')
@login_required
def list_orders():
    tenant_id = current_user.tenant_id
    tab = request.args.get('tab', '')
    page = request.args.get('page', 1, type=int)

    orders = service.get_orders(tenant_id, tab=tab or None, page=page)
    stats = service.get_order_stats(tenant_id)

    return render_template('products/orders.html', orders=orders, tab=tab,
                           order_types=ORDER_TYPES, stats=stats)


@products_bp.route('/orders/create', methods=['GET', 'POST'])
@login_required
def create_order():
    order_type = request.args.get('type', 'planned_import')

    if request.method == 'POST':
        order = service.create_order(current_user.tenant_id, request.form, current_user.id)
        flash(f'Đã tạo phiếu {order.order_number}!', 'success')
        return redirect(url_for('products.order_detail', id=order.id))

    products = service.get_active_physical_products(current_user.tenant_id)
    return render_template('products/order_form.html', order=None,
                           order_type=order_type, order_types=ORDER_TYPES,
                           products=products)


@products_bp.route('/orders/<int:id>')
@login_required
def order_detail(id):
    order = service.get_order(current_user.tenant_id, id)
    return render_template('products/order_detail.html', order=order,
                           order_types=ORDER_TYPES, order_statuses=ORDER_STATUSES)


@products_bp.route('/orders/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_order(id):
    order = service.get_order(current_user.tenant_id, id)
    if order.status != 'draft':
        flash('Chỉ sửa được phiếu ở trạng thái Nháp.', 'error')
        return redirect(url_for('products.order_detail', id=order.id))

    if request.method == 'POST':
        service.update_order(current_user.tenant_id, id, request.form)
        flash('Đã cập nhật phiếu!', 'success')
        return redirect(url_for('products.order_detail', id=order.id))

    products = service.get_active_physical_products(current_user.tenant_id)
    return render_template('products/order_form.html', order=order,
                           order_type=order.type, order_types=ORDER_TYPES,
                           products=products)


@products_bp.route('/orders/<int:id>/complete', methods=['POST'])
@login_required
def complete_order(id):
    order, error = service.complete_order(current_user.tenant_id, id, current_user.id)
    if error:
        flash(error, 'error')
    else:
        flash(f'Phiếu {order.order_number} đã hoàn thành! Tồn kho đã cập nhật.', 'success')
    return redirect(url_for('products.order_detail', id=id))


@products_bp.route('/orders/<int:id>/cancel', methods=['POST'])
@login_required
def cancel_order(id):
    order, error = service.cancel_order(current_user.tenant_id, id)
    if error:
        flash(error, 'error')
    else:
        flash(f'Đã hủy phiếu {order.order_number}.', 'success')
    return redirect(url_for('products.order_detail', id=id))


@products_bp.route('/orders/<int:id>/delete', methods=['POST'])
@login_required
def delete_order(id):
    success, error = service.delete_order(current_user.tenant_id, id)
    if error:
        flash(error, 'error')
        return redirect(url_for('products.order_detail', id=id))
    flash('Đã xóa phiếu.', 'success')
    return redirect(url_for('products.list_orders'))


# ============================================================
# KIỂM KHO HÀNG LOẠT (InventoryCount)
# ============================================================
@products_bp.route('/inventory-count')
@login_required
def list_counts():
    page = request.args.get('page', 1, type=int)
    counts = service.get_counts(current_user.tenant_id, page=page)
    return render_template('products/count_list.html', counts=counts)


@products_bp.route('/inventory-count/create', methods=['GET', 'POST'])
@login_required
def create_count():
    if request.method == 'POST':
        count = service.create_count(current_user.tenant_id, request.form, current_user.id)
        flash(f'Phiếu kiểm kho {count.count_number} — {count.total_products} sản phẩm', 'success')
        return redirect(url_for('products.count_detail', id=count.id))

    products = service.get_active_physical_products(current_user.tenant_id)
    return render_template('products/count_create.html', products=products)


@products_bp.route('/inventory-count/<int:id>', methods=['GET', 'POST'])
@login_required
def count_detail(id):
    count = service.get_count(current_user.tenant_id, id)

    if request.method == 'POST' and count.status == 'draft':
        service.update_count_items(current_user.tenant_id, id, request.form)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'ok', 'message': 'Đã lưu'})

        flash('Đã cập nhật số liệu kiểm kho!', 'success')
        return redirect(url_for('products.count_detail', id=id))

    return render_template('products/count_detail.html', count=count)


@products_bp.route('/inventory-count/<int:id>/complete', methods=['POST'])
@login_required
def complete_count(id):
    count, result = service.complete_count(current_user.tenant_id, id, current_user.id)
    if count is None:
        flash(result, 'error')
    else:
        flash(f'Kiểm kho hoàn thành! Đã điều chỉnh {result} sản phẩm.', 'success')
    return redirect(url_for('products.count_detail', id=id))


@products_bp.route('/inventory-count/<int:id>/cancel', methods=['POST'])
@login_required
def cancel_count(id):
    count, error = service.cancel_count(current_user.tenant_id, id)
    if error:
        flash(error, 'error')
    else:
        flash('Đã hủy phiếu kiểm kho.', 'success')
    return redirect(url_for('products.count_detail', id=id))


# ============================================================
# EXCEL IMPORT / EXPORT
# ============================================================
@products_bp.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Vui lòng chọn file Excel (.xlsx)', 'error')
            return redirect(url_for('products.import_excel'))

        try:
            import openpyxl
            from models.user import db
            from models.product import Product, StockTransaction

            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            imported = 0
            updated = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:
                    continue
                try:
                    name = str(row[0]).strip()
                    sku = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    category = str(row[2]).strip() if len(row) > 2 and row[2] else 'Khác'
                    unit = str(row[3]).strip() if len(row) > 3 and row[3] else 'Cái'
                    sell_price = float(row[4]) if len(row) > 4 and row[4] else 0
                    cost_price = float(row[5]) if len(row) > 5 and row[5] else 0
                    stock_qty = float(row[6]) if len(row) > 6 and row[6] else 0
                    min_stock = float(row[7]) if len(row) > 7 and row[7] else 0
                    description = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                    barcode = str(row[9]).strip() if len(row) > 9 and row[9] else ''

                    existing = None
                    if sku:
                        existing = Product.query.filter_by(
                            tenant_id=current_user.tenant_id, sku=sku).first()

                    if existing:
                        existing.name = name
                        existing.category = category if category in PRODUCT_CATEGORIES else existing.category
                        existing.unit = unit if unit in PRODUCT_UNITS else existing.unit
                        existing.sell_price = sell_price
                        existing.cost_price = cost_price
                        existing.min_stock = min_stock
                        if description:
                            existing.description = description
                        if barcode:
                            existing.barcode = barcode

                        if stock_qty != existing.stock_quantity:
                            old_qty = existing.stock_quantity
                            existing.stock_quantity = stock_qty
                            db.session.add(StockTransaction(
                                tenant_id=current_user.tenant_id, product_id=existing.id,
                                type='adjust', quantity=stock_qty - old_qty,
                                unit_cost=cost_price, stock_before=old_qty, stock_after=stock_qty,
                                reference='Excel Import', note='Cập nhật từ Excel',
                                created_by=current_user.id,
                            ))
                        updated += 1
                    else:
                        product = Product(
                            tenant_id=current_user.tenant_id, created_by=current_user.id,
                            owner_id=current_user.id, name=name, sku=sku, barcode=barcode,
                            category=category if category in PRODUCT_CATEGORIES else 'Khác',
                            unit=unit if unit in PRODUCT_UNITS else 'Cái',
                            description=description, sell_price=sell_price,
                            cost_price=cost_price, stock_quantity=stock_qty,
                            min_stock=min_stock, is_active=True,
                        )
                        db.session.add(product)
                        db.session.flush()
                        if stock_qty > 0:
                            db.session.add(StockTransaction(
                                tenant_id=current_user.tenant_id, product_id=product.id,
                                type='import', quantity=stock_qty, unit_cost=cost_price,
                                stock_before=0, stock_after=stock_qty,
                                reference='Excel Import', note='Nhập từ Excel',
                                created_by=current_user.id,
                            ))
                        imported += 1
                except Exception as e:
                    errors.append(f'Dòng {row_idx}: {str(e)}')

            db.session.commit()
            msg = f'Import thành công: {imported} mới, {updated} cập nhật.'
            if errors:
                msg += f' Lỗi: {len(errors)} dòng.'
            flash(msg, 'success' if not errors else 'warning')
        except Exception as e:
            flash(f'Lỗi đọc file: {str(e)}', 'error')

        return redirect(url_for('products.list_products'))

    return render_template('products/import.html')


@products_bp.route('/export')
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from models.product import Product

    products = Product.query.filter_by(tenant_id=current_user.tenant_id)\
        .order_by(Product.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sản phẩm'

    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='4976F4', end_color='4976F4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Tên sản phẩm', 'SKU', 'Danh mục', 'Đơn vị', 'Giá bán',
               'Giá vốn', 'Tồn kho', 'Tồn tối thiểu', 'Mô tả', 'Barcode']

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = halign
        c.border = border

    for ri, p in enumerate(products, 2):
        data = [p.name, p.sku or '', p.category, p.unit, p.sell_price,
                p.cost_price, p.stock_quantity, p.min_stock, p.description or '', p.barcode or '']
        for col, val in enumerate(data, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.border = border

    for col in ws.columns:
        ml = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'san_pham_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')


@products_bp.route('/template')
@login_required
def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Nhập sản phẩm'

    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='4976F4', end_color='4976F4', fill_type='solid')
    nf = Font(italic=True, color='8C8C8C', size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Tên sản phẩm *', 'SKU', 'Danh mục', 'Đơn vị', 'Giá bán',
               'Giá vốn', 'Tồn kho', 'Tồn tối thiểu', 'Mô tả', 'Barcode']
    examples = ['Phần mềm CRM Pro', 'SP-001', 'Phần mềm', 'License', 50000000,
                30000000, 100, 10, 'Phần mềm quản lý', '8901234567890']

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal='center')
        c.border = border

    for col, val in enumerate(examples, 1):
        c = ws.cell(row=2, column=col, value=val)
        c.font = nf
        c.border = border

    ws_notes = wb.create_sheet('Hướng dẫn')
    for r, row_data in enumerate([
        ['HƯỚNG DẪN IMPORT'], [''],
        ['1. Cột "Tên sản phẩm" bắt buộc (*)'], ['2. SKU trùng → cập nhật SP hiện có'],
        ['3. SKU mới/trống → tạo SP mới'], [''],
        ['DANH MỤC:'], [', '.join(PRODUCT_CATEGORIES)], [''],
        ['ĐƠN VỊ:'], [', '.join(PRODUCT_UNITS)],
    ], 1):
        for c, val in enumerate(row_data, 1):
            ws_notes.cell(row=r, column=c, value=val)
    ws_notes['A1'].font = Font(bold=True, size=14)
    ws_notes.column_dimensions['A'].width = 60

    widths = [25, 12, 15, 10, 15, 15, 10, 12, 30, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='mau_import_san_pham.xlsx')


# ============================================================
# API
# ============================================================
@products_bp.route('/api/search')
@login_required
def api_search():
    q = request.args.get('q', '').strip()
    results = service.api_search(current_user.tenant_id, q=q)
    return jsonify(results)
